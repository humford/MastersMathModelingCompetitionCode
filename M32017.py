# M3 2017 Model Calculation
# Team 9081

import math
import numpy
from scipy import integrate, stats, interpolate
from openpyxl import load_workbook

NPS_airqual = load_workbook('NPS_Air_Quality_0.xlsx', read_only = True)      
NPS_heatindex = load_workbook('NPS_Heat_Index_0.xlsx', read_only = True)      
NPS_hurricanes = load_workbook('NPS_Hurricanes_0.xlsx', read_only = True)      
NPS_sealevel = load_workbook('NPS_MeanSeaLevel.xlsx', read_only = True)      
NPS_temperature = load_workbook('NPS_Temperature.xlsx', read_only = True)      
NPS_visitors = load_workbook('NPS_Visitor_Stats_0.xlsx', read_only = True)      
NPS_wildfires = load_workbook('NPS_Wildfires_1.xlsx', read_only = True)      

def aq(unit, t):
	if unit == 3:
		ws = NPS_airqual["Acadia NP"]
	if unit == 4:
		ws = NPS_airqual["Cape Hatteras NS"]
	if unit == 5:
		ws = NPS_airqual["Kenai Fjords NP"]
	if unit == 6:
		ws = NPS_airqual["Olympic NP"]
	if unit == 7:
		ws = NPS_airqual["Padre Island NS"]
	months = []
	airqual = []
	airqualavg = []
	month = 1
	for row in range(5, 257): 
		if ws.cell(column=3, row=row).value != "Missing": 
			airqualavg.append(float(ws.cell(column=3, row=row).value))
	avg = sum(airqualavg) / len(airqualavg)
	for row in range(5, 257):
		if ws.cell(column=3, row=row).value != "Missing":
			airqual.append(float(ws.cell(column=3, row=row).value))
		else:
			airqual.append(avg)
		months.append(month)
		month += 1
	p = stats.linregress(months, airqual)
	return airqual[-1] - (p[0] * ((t*12) + 252) + p[1])

def hu(unit, t):
	if unit == 3:
		ws = NPS_hurricanes["Acadia NP"]
		top = 25
	if unit == 4:
		ws = NPS_hurricanes["Cape Hatteras NS"]
		top = 38
	if unit == 5:
		ws = NPS_hurricanes["Kenai Fjords NP"]
		top = 25
	if unit == 6:
		ws = NPS_hurricanes["Olympic NP"]
		top = 25
	if unit == 7:
		ws = NPS_hurricanes["Padre Island NS"]
		top = 27
	hurricanes = {}
	for row in range(4, top):
		if ws.cell(column=1, row=row).value != None:
			if hurricanes.get(int(ws.cell(column=1, row=row).value)) == None:
				hurricanes[int(ws.cell(column=1, row=row).value)] = [ws.cell(column=3, row=row).value]
			else:
				hurricanes[int(ws.cell(column=1, row=row).value)].append(ws.cell(column=3, row=row).value)
	hurricanesavg = []
	for y in range(1, 21):
		if hurricanes.get(1996+y) != None:
			s = 0
			for i in hurricanes[1996+y]:
				#ET =3.2 10-4 
				#TD=.2502
				#TS=.5002
				#H1=.7500
				#H2=1
				#H3=1.2500
				#H4=1.5000
				#H5=1.7500
				if i == "ET":
					s += 0.00032
				elif i == "TD":
					s += 0.2502
				elif i == "TS":
					s += 0.5002
				elif i == "H1":
					s += 0.7500
				elif i == "H2":
					s += 1.0000
				elif i == "H3":
					s += 1.2500
				elif i == "H4":
					s += 1.5000
				elif i == "H5":
					s += 1.7500
			hurricanesavg.append(s)
	avg = sum(hurricanesavg)/len(hurricanesavg)
	years = []
	hu = []
	for y in range(1, 21):
		years.append(y)
		if hurricanes.get(1996+y) != None:
			s = 0
			for i in hurricanes[1996+y]:
				#ET =3.2 10-4 
				#TD=.2502
				#TS=.5002
				#H1=.7500
				#H2=1
				#H3=1.2500
				#H4=1.5000
				#H5=1.7500
				if i == "ET":
					s += 0.00032
				elif i == "TD":
					s += 0.2502
				elif i == "TS":
					s += 0.5002
				elif i == "H1":
					s += 0.7500
				elif i == "H2":
					s += 1.0000
				elif i == "H3":
					s += 1.2500
				elif i == "H4":
					s += 1.5000
				elif i == "H5":
					s += 1.7500
			hu.append(s)
		else:
			hu.append(avg)
	p = numpy.polyfit(numpy.array(years), numpy.array(hu), 2)
	i = integrate.quad(lambda x: p[0] * x**2 + p[1] * x + p[2], 21, t + 21) 
	return i[0]
	#p = numpy.polyfit(numpy.array(years), numpy.array(hu), 3)
	#i = integrate.quad(lambda x: p[0] * x**3 + p[1] * x**2 + p[2] * x + p[3], 21, t + 21) 
	#return i[0]

def hi(unit, t):
	if unit == 3:
		ws = NPS_heatindex["Acadia NP"]
	if unit == 4:
		ws = NPS_heatindex["Cape Hatteras NS"]
	if unit == 5:
		ws = NPS_heatindex["Kenai Fjords NP"]
	if unit == 6:
		ws = NPS_heatindex["Olympic NP"]
	if unit == 7:
		ws = NPS_heatindex["Padre Island NS"]
	months = []
	heatindex = []
	heatindexavg = []
	month = 1
	for row in range(4, 256): 
		if ws.cell(column=3, row=row).value != "Missing": 
			heatindexavg.append(float(ws.cell(column=3, row=row).value))
	avg = sum(heatindexavg) / len(heatindexavg)
	for row in range(4, 256):
		if ws.cell(column=3, row=row).value != "Missing":
			heatindex.append(float(ws.cell(column=3, row=row).value))
		else:
			heatindex.append(avg)
		months.append(month)
		month += 1
	p = stats.linregress(months, airqual)
	return p[0] * ((t*12) + 252) + p[1]

def sl(unit, t):
	ws = NPS_sealevel["20yr_Monthly_MSL"]
	months = []
	sealevels = []
	sealevelsavg = []
	month = 1
	for row in range(5, 245): 
		if ws.cell(column=unit, row=row).value != " ": 
			sealevelsavg.append(float(ws.cell(column=unit, row=row).value))
	avg = sum(sealevelsavg) / len(sealevelsavg)
	for row in range(5, 245):
		if ws.cell(column=unit, row=row).value != " ":
			sealevels.append(float(ws.cell(column=unit, row=row).value))
		else:
			sealevels.append(avg)
		months.append(month)
		month += 1
	p = numpy.polyfit(numpy.array(months), numpy.array(sealevels), 2)
	i = integrate.quad(lambda x: p[0] * x**2 + p[1] * x + p[2], 241, (t * 12) + 241) 
	return i[0]

def te(unit, t):
	if unit == 3:
		ws = NPS_temperature["Acadia NP"]
	if unit == 4:
		ws = NPS_temperature["Cape Hatteras NS"]
	if unit == 5:
		ws = NPS_temperature["Kenai Fjords NP"]
	if unit == 6:
		ws = NPS_temperature["Olympic NP"]
	if unit == 7:
		ws = NPS_temperature["Padre Island NS"]
	months = []
	temp = []
	tempavg = []
	month = 1
	for row in range(4, 210): 
		if ws.cell(column=10, row=row).value != None: 
			tempavg.append(float(float((ws.cell(column=10, row=row).value) - 32.0) * (5.0/9.0)))
	avg = sum(tempavg) / len(tempavg)
	for row in range(4, 210):
		if ws.cell(column=10, row=row).value != None:
			temp.append(float(float((ws.cell(column=10, row=row).value) - 32.0) * (5.0/9.0)))
		else:
			temp.append(avg)
		months.append(month)
		month += 1
	p = stats.linregress(months, temp)
	return (p[0] * ((t*12) + 206) + p[1]) - temp[-1]

def wi(unit, t):
	if unit == 3:
		ws = NPS_wildfires["Acadia NP"]
		top = 75
	if unit == 4:
		return 19681.29619503301
		ws = NPS_wildfires["Cape Hatteras NS"]
		top = 88
	if unit == 5:
		return 0
	if unit == 6:
		ws = NPS_wildfires["Olympic NP"]
		top = 43
	if unit == 7:
		ws = NPS_wildfires["Padre Island NS"]
		top = 405
	wildfires = {}
	for row in range(2, top):
		if ws.cell(column=1, row=row).value != None:
			if wildfires.get(int(ws.cell(column=1, row=row).value)) == None:
				wildfires[int(ws.cell(column=1, row=row).value)] = float(ws.cell(column=12, row=row).value)
			else:
				wildfires[int(ws.cell(column=1, row=row).value)] += float(ws.cell(column=12, row=row).value)
	years = []
	wf = []
	for y in range(1, 21):
		years.append(y)
		if wildfires.get(1996+y) != None:
			wf.append(wildfires[1996+y])
		else:
			wf.append(sum(wildfires.values())/len(wildfires))
	p = numpy.polyfit(numpy.array(years), numpy.array(wf), 3)
	i = integrate.quad(lambda x: p[0] * x**3 + p[1] * x**2 + p[2] * x + p[3], 21, t + 21) 
	return i[0]

def vs(unit, t):
	if unit == 3:
		ws = NPS_visitors["Acadia NP"]
		top = 39
	if unit == 4:
		ws = NPS_visitors["Cape Hatteras NS"]
		top = 39
	if unit == 5:
		ws = NPS_visitors["Kenai Fjords NP"]
		top = 36
	if unit == 6:
		ws = NPS_visitors["Olympic NP"]
		top = 39
	if unit == 7:
		ws = NPS_visitors["Padre Island NS"]
		top = 39
	years = []
	visitors = []
	year = 1
	for row in range(top,3,-1):
		if ws.cell(column=unit, row=row).value != " ":
			visitors.append(int(ws.cell(column=unit, row=row).value))
		years.append(year)
		year += 1
	p = stats.linregress(years, visitors)
	i = integrate.quad(lambda x: p[0] * x + p[1], top - 19, t + (top - 19)) 
	return i[0] 

def h(unit):
	return (1.1 ** (sl(unit, 50)/50)) * hu(unit, 50) - (hu(unit, 50)/50)

def w(unit):
	return wi(unit, 50) + 0.512 * te(unit, 50)

def V(unit):
	return 3512642 * h(unit) + 128 * w(unit) + 143141 * aq(unit, 50)

def R(unit):
	if unit == 3:
		cost = 16.0
	if unit == 4:
		cost = 13.5
	if unit == 5:
		cost = 8.0
	if unit == 6:
		cost = 0.0
	if unit == 7:
		cost = 7.5
	return vs(unit, 50) * cost

def I(unit):
	return R(unit) - V(unit)

print R(3)
print R(4)
print R(5)
print R(6)
print R(7)

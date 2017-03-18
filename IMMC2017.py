# IMMC 2017 Model Calculation
# Team # US-6462

from math import radians, cos, sin, asin, sqrt
# Distance between two lat/lng coordinates in km using the Haversine formula
def getDistanceFromLatLng(lat1, lng1, lat2, lng2, miles=False): # use decimal degrees
  r=6371 # radius of the earth in km
  lat1=radians(lat1)
  lat2=radians(lat2)
  lat_dif=lat2-lat1
  lng_dif=radians(lng2-lng1)
  a=sin(lat_dif/2.0)**2+cos(lat1)*cos(lat2)*sin(lng_dif/2.0)**2
  d=2*r*asin(sqrt(a))
  if miles:
    return d * 0.621371 # return miles
  else:
    return d # return km
# Copyright 2016, Chris Youderian, SimpleMaps, http://simplemaps.com/resources/location-distance
# Released under MIT license - https://opensource.org/licenses/MIT

import wbpy
import easygui

import operator
import re

import json
import requests
from urllib import urlopen

import googlemaps
import astral
from datetime import datetime, timedelta
from time import sleep
from dateutil import parser
from timezonefinder import TimezoneFinder
from pytz import timezone, utc

from openpyxl import load_workbook
cities = {u'Kinshasa': {'lat': -4.32758, 'lng': 15.31357}, u'Zhengzhou': {'lat': 34.75778, 'lng': 113.64861}, u'Rio de Janeiro': {'lat': -22.90278, 'lng': -43.2075}, u'Zhongshan': {'lat': 21.31992, 'lng': 110.5723}, u'Sao Paulo': {'lat': -23.5475, 'lng': -46.63611}, u'Shenzhen': {'lat': 22.54554, 'lng': 114.0683}, u'Wuhan': {'lat': 30.58333, 'lng': 114.26667}, u'Santiago': {'lat': -33.45694, 'lng': -70.64827}, u'Ankara': {'lat': 39.91987, 'lng': 32.85427}, u'Shiyan': {'lat': 32.6475, 'lng': 110.77806}, u'Delhi': {'lat': 28.65195, 'lng': 77.23149}, u"Xi'an": {'lat': 34.25833, 'lng': 108.92861}, u'Tianshui': {'lat': 34.57952, 'lng': 105.74238}, u'Baghdad': {'lat': 33.34058, 'lng': 44.40088}, u'Taipei': {'lat': 25.04776, 'lng': 121.53185}, u'Ahmedabad': {'lat': 23.02579, 'lng': 72.58727}, u'Jeddah': {'lat': 21.54238, 'lng': 39.19797}, u'Pune': {'lat': 18.51957, 'lng': 73.85535}, u'Chittagong': {'lat': 22.3384, 'lng': 91.83168}, u'Tangshan': {'lat': 39.63333, 'lng': 118.18333}, u'Surat': {'lat': 21.19594, 'lng': 72.83023}, u'Harbin': {'lat': 45.75, 'lng': 126.65}, u'Kanpur': {'lat': 26.46523, 'lng': 80.34975}, u'Alexandria': {'lat': 31.21564, 'lng': 29.95527}, u'Casablanca': {'lat': 33.58831, 'lng': -7.61138}, u'Cairo': {'lat': 30.06263, 'lng': 31.24967}, u'Shenyang': {'lat': 41.79222, 'lng': 123.43278}, u'Abidjan': {'lat': 5.30966, 'lng': -4.01266}, u'London': {'lat': 51.50853, 'lng': -0.12574}, u'Caracas': {'lat': 10.48801, 'lng': -66.87919}, u'Puyang': {'lat': 29.45679, 'lng': 119.88872}, u'Zibo': {'lat': 36.79056, 'lng': 118.06333}, u'Sydney': {'lat': -33.86785, 'lng': 151.20732}, u'Ho Chi Minh City': {'lat': 10.82302, 'lng': 106.62965}, u'UEruemqi': {'lat': 43.80096, 'lng': 87.60046}, u'Seoul': {'lat': 37.566, 'lng': 126.9784}, u'Singapore': {'lat': 1.28967, 'lng': 103.85007}, u'New York City': {'lat': 40.71427, 'lng': -74.00597}, u'Moscow': {'lat': 55.75222, 'lng': 37.61556}, u'Guangzhou': {'lat': 23.11667, 'lng': 113.25}, u'Hefei': {'lat': 31.86389, 'lng': 117.28083}, u'Berlin': {'lat': 52.52437, 'lng': 13.41053}, u'Los Angeles': {'lat': 34.05223, 'lng': -118.24368}, u'Tehran': {'lat': 35.69439, 'lng': 51.42151}, u'Karachi': {'lat': 24.9056, 'lng': 67.0822}, u'Xiamen': {'lat': 24.47979, 'lng': 118.08187}, u'Ibadan': {'lat': 7.37756, 'lng': 3.90591}, u"Tai'an": {'lat': 36.18528, 'lng': 117.12}, u'Kano': {'lat': 12.00012, 'lng': 8.51672}, u'Cape Town': {'lat': -33.92584, 'lng': 18.42322}, u'Foshan': {'lat': 23.02677, 'lng': 113.13148}, u'Hangzhou': {'lat': 30.29365, 'lng': 120.16142}, u'Chennai': {'lat': 13.08784, 'lng': 80.27847}, u'Wuxi': {'lat': 31.56887, 'lng': 120.28857}, u'Taiyuan': {'lat': 37.86944, 'lng': 112.56028}, u'Shijiazhuang': {'lat': 38.04139, 'lng': 114.47861}, u'Hyderabad': {'lat': 17.38405, 'lng': 78.45636}, u'Madrid': {'lat': 40.4165, 'lng': -3.70256}, u'Busan': {'lat': 35.10278, 'lng': 129.04028}, u'Ningbo': {'lat': 29.87819, 'lng': 121.54945}, u'Mexico City': {'lat': 19.42847, 'lng': -99.12766}, u'Istanbul': {'lat': 41.01384, 'lng': 28.94966}, u'Bengaluru': {'lat': 12.97194, 'lng': 77.59369}, u'Yangon': {'lat': 16.80528, 'lng': 96.15611}, u'Lahore': {'lat': 31.54972, 'lng': 74.34361}, u'Dongguan': {'lat': 23.01797, 'lng': 113.74866}, u'Shantou': {'lat': 23.36814, 'lng': 116.71479}, u'Durban': {'lat': -29.8579, 'lng': 31.0292}, u'Kunming': {'lat': 25.03889, 'lng': 102.71833}, u'Hong Kong': {'lat': 22.28552, 'lng': 114.15769}, u'Tokyo': {'lat': 35.6895, 'lng': 139.69171}, u'Nanchong': {'lat': 30.79508, 'lng': 106.08473}, u'Beijing': {'lat': 39.9075, 'lng': 116.39723}, u'Lima': {'lat': -12.04318, 'lng': -77.02824}, u'Mumbai': {'lat': 19.07283, 'lng': 72.88261}, u'Tianjin': {'lat': 39.14222, 'lng': 117.17667}, u'Shanghai': {'lat': 31.22222, 'lng': 121.45806}, u'Kabul': {'lat': 34.52813, 'lng': 69.17233}, u'Kolkata': {'lat': 22.56263, 'lng': 88.36304}, u'Saint Petersburg': {'lat': 59.93863, 'lng': 30.31413}, u'City of London': {'lat': 51.51279, 'lng': -0.09184}, u'Buenos Aires': {'lat': -34.61315, 'lng': -58.37723}, u'Changchun': {'lat': 43.88, 'lng': 125.32278}, u'Bangkok': {'lat': 13.75398, 'lng': 100.50144}, u'Qingdao': {'lat': 36.06605, 'lng': 120.36939}, u'Chongqing': {'lat': 29.56278, 'lng': 106.55278}, u'Dhaka': {'lat': 23.7104, 'lng': 90.40744}, u'Jakarta': {'lat': -6.21462, 'lng': 106.84513}, u'Changsha': {'lat': 28.19874, 'lng': 112.97087}, u'Melbourne': {'lat': -37.814, 'lng': 144.96332}, u'Suzhou': {'lat': 31.30408, 'lng': 120.59538}, u'Dalian': {'lat': 38.91222, 'lng': 121.60222}, u'Chengdu': {'lat': 30.66667, 'lng': 104.06667}, u'Riyadh': {'lat': 24.68773, 'lng': 46.72185}, u'Bogota': {'lat': 4.60971, 'lng': -74.08175}, u'Jinan': {'lat': 36.66833, 'lng': 116.99722}, u'Nanjing': {'lat': 32.06167, 'lng': 118.77778}, u'Lagos': {'lat': 6.45407, 'lng': 3.39467}, u'Yokohama': {'lat': 35.43333, 'lng': 139.65}, u'Pyongyang': {'lat': 39.03385, 'lng': 125.75432}}

gmaps = googlemaps.Client(key='AIzaSyBwml6CdKQRbEXICEOMoUsc_iBVB46Nlyc') #gmaps API key
a = astral.Astral()
a.solar_depression = 'civil'
tf = TimezoneFinder()
api = 0
location_table = {}

def lost_productivity_city_participant(time, participant, city):
	l = astral.Location()
	l.latitude = city[0]
	l.longitude = city[1]
	Ts = l.daylight(date=time, local=False)
	Tb = utc.localize(time + timedelta(seconds= 32400))
	Te = utc.localize(time + timedelta(seconds= 61200))
	if Ts[0] <= Tb and Ts[1] >= Te:
		Tl = 8.0
	elif Ts[0] > Tb and Ts[1] >= Te:
		Tl = 8.0 - ((Ts[0] - Tb).seconds // 3600)
	elif Ts[0] <= Tb and Ts[1] < Te:
		Tl = 8.0 - ((Te - Ts[1]).seconds // 3600)
	elif Ts[0] > Tb and Ts[1] < Te:
		Tl = ((Ts[1] - Ts[0]).seconds // 3600)

	target_datetime = time.replace(tzinfo= timezone(tf.timezone_at(lat= city[0], lng= city[1])) )
	participant_datetime = time.replace(tzinfo= timezone(tf.timezone_at(lat= location_table[participant[0]]["lat"], lng= location_table[participant[0]]["lng"])) )
	tz_crossed = target_datetime.astimezone(utc) - participant_datetime.astimezone(utc)
	if participant_datetime.astimezone(utc) > target_datetime.astimezone(utc):
		dr1 = round(float(abs(tz_crossed.seconds // 3600) / 1.53))
		dr2 = round(float( (abs(tz_crossed.seconds // 3600) - 1.53) / ((92.0+5.75*Tl)/60.0) ))
		C = ((0.07 * (92.0+5.75*Tl)) / 60.0) + 1.0
		psum = 0
		for d in range(1, 4):
			if d == 1:
				psum += (1.0/(1.10733333333 ** dr1)) * (1.10733333333 ** d)
			else:
				if d > dr2:
					d = dr2
				psum += (1.0/(C ** dr2)) * (C ** d)
		p = psum / 3
		return (1.0-p) * ((participant[1] / 250) * 3)
	elif participant_datetime.astimezone(utc) < target_datetime.astimezone(utc):
		dr1 = round(float(abs(tz_crossed.seconds // 3600) / 0.95))
		dr2 = round(float( (abs(tz_crossed.seconds // 3600) - 1.53) / ((57.0+5.75*Tl)/60.0) ))
		C = ((0.07 * (57.0+5.75*Tl)) / 60.0) + 1.0
		psum = 0
		for d in range(1, 4):
			if d == 1:
				psum += (1.0/(1.0665 ** dr1)) * (1.0665 ** d)
			else:
				if d > dr2:
					d = dr2
				psum += (1.0/(C ** dr2)) * (C ** d)
		p = psum / 3
		return (1.0-p) * ((participant[1] / 250) * 3)
	else:
		return 0

def lost_productivity_city(time, participants, city):
	lost = 0
	for participant in participants:
		lost += lost_productivity_city_participant(time, [participant, participants[participant]], city)
	return lost 

def lost_productivity(time, participants):
	smallest = 10000000000
	smallest_city = "a"
	city_prices = {}
	for city in cities:
		lpc = lost_productivity_city(time, participants, [ cities[city]["lat"], cities[city]["lng"] ])
		city_prices[city] = round(lpc, 2)
	
	ws = output["SmallMeetingProductivity"]
	sorted_city_prices = sorted(city_prices.items(), key=operator.itemgetter(1))
	for x in range(0, len(sorted_city_prices)): 
		ws['A' + str(x+1)] = sorted_city_prices[x][0]
		ws['B' + str(x+1)] = sorted_city_prices[x][1]
	print "Productivity"

	return city_prices

def cost_city_participant(time, participant, city):
	if time < datetime.now():
		time += timedelta(seconds=31536000)
	global api
	if api == 0:
		url = "http://partners.api.skyscanner.net/apiservices/browsequotes/v1.0/US/USD/en-US/" + str(location_table[participant[0]]["lat"]) + "," + str(location_table[participant[0]]["lng"]) + "-latlong/" + str(city[0]) + "," + str(city[1]) + "-latlong/" + str(time.strftime("%Y-%m")) + "/" + str(time.strftime("%Y-%m")) + "?apikey=prtl6749387986743898559646983194"
		api = 1
	elif api == 1:
		url = "http://partners.api.skyscanner.net/apiservices/browsequotes/v1.0/US/USD/en-US/" + str(location_table[participant[0]]["lat"]) + "," + str(location_table[participant[0]]["lng"]) + "-latlong/" + str(city[0]) + "," + str(city[1]) + "-latlong/" + str(time.strftime("%Y-%m")) + "/" + str(time.strftime("%Y-%m")) + "?apikey=ma403936798903192832036949206853"
		api = 2
	elif api == 2:
		url = "http://partners.api.skyscanner.net/apiservices/browsequotes/v1.0/US/USD/en-US/" + str(location_table[participant[0]]["lat"]) + "," + str(location_table[participant[0]]["lng"]) + "-latlong/" + str(city[0]) + "," + str(city[1]) + "-latlong/" + str(time.strftime("%Y-%m")) + "/" + str(time.strftime("%Y-%m")) + "?apikey=bi648589294070541234977848866042"
		api = 3
	elif api == 3:
		url = "http://partners.api.skyscanner.net/apiservices/browsequotes/v1.0/US/USD/en-US/" + str(location_table[participant[0]]["lat"]) + "," + str(location_table[participant[0]]["lng"]) + "-latlong/" + str(city[0]) + "," + str(city[1]) + "-latlong/" + str(time.strftime("%Y-%m")) + "/" + str(time.strftime("%Y-%m")) + "?apikey=ja674195465513836518072649821928"
		api = 4
	elif api == 4:
		url = "http://partners.api.skyscanner.net/apiservices/browsequotes/v1.0/US/USD/en-US/" + str(location_table[participant[0]]["lat"]) + "," + str(location_table[participant[0]]["lng"]) + "-latlong/" + str(city[0]) + "," + str(city[1]) + "-latlong/" + str(time.strftime("%Y-%m")) + "/" + str(time.strftime("%Y-%m")) + "?apikey=ma355026595362938660434727273431"
		api = 5
	elif api == 5:
		url = "http://partners.api.skyscanner.net/apiservices/browsequotes/v1.0/US/USD/en-US/" + str(location_table[participant[0]]["lat"]) + "," + str(location_table[participant[0]]["lng"]) + "-latlong/" + str(city[0]) + "," + str(city[1]) + "-latlong/" + str(time.strftime("%Y-%m")) + "/" + str(time.strftime("%Y-%m")) + "?apikey=ma842007923172588486539661677492"
		api = 0
	json_content = urlopen(url).read()
	sleep(1.5)
	print json_content
	count = 1
	while json_content == "{\"ValidationErrors\":[{\"Message\":\"Rate limit has been exceeded: 20 PerMinute for Browse\"}]}":
		print "ValError" + str(count)
		json_content = urlopen(url).read()
		sleep(10)
		count += 1
	content = json.loads(json_content)
	if "Quotes" in content.keys():
		if content["Quotes"]:
			fares = []
			for quote in content["Quotes"]:
				fares.append(quote["MinPrice"])
			fare = min(fares)
		else:
			fare = 2 * (50 + (0.11 * getDistanceFromLatLng(location_table[participant[0]]["lat"], location_table[participant[0]]["lng"], city[0], city[1], True)))
			#Fare = $50 + (Distance * $0.11)
	return fare

def cost_city(time, participants, city):
	cost = 0
	for participant in participants:
		ccp = cost_city_participant(time, [participant, participants[participant]], city)
		cost += ccp
	return cost

def cost(time, participants):
	city_prices = {}
	for city in cities:
		cc = cost_city(time, participants, [ cities[city]["lat"], cities[city]["lng"] ])
		if cc != 0:
			city_prices[city] = round(cc, 2)
	
	ws = output["SmallMeetingCost"]
	sorted_city_prices = sorted(city_prices.items(), key=operator.itemgetter(1))
	for x in range(0, len(sorted_city_prices)): 
		ws['A' + str(x+1)] = sorted_city_prices[x][0]
		ws['B' + str(x+1)] = sorted_city_prices[x][1]
	print "Cost Written"

	return city_prices

def best_meeting_location(time, participants):
	lp = lost_productivity(time, participants)
	c = cost(time, participants)
	bml = {}
	for city in lp:
		bml[city] = [lp[city],c[city],lp[city] + c[city]]
	sorted_city_prices = sorted(bml.items(), key=operator.itemgetter(1))
	sorted_city_prices = sorted_city_prices[0:5]
	return sorted_city_prices

def gui():
	easygui.msgbox('Find a place to hold your meeting. ~Team # US-6462', 'International Meeting Management Corporation')
	msg = "Enter the month of your meeting and the number of participants"
	title = "International Meeting Management Corporation"
	fieldNames = ["Time","Participants"]
	fieldValues = []  # we start with blanks for the values
	fieldValues = easygui.multenterbox(msg,title,fieldNames)
	# make sure that none of the fields was left blank
	while 1:
		if fieldValues == None: 
			break
		errmsg = ""
		for i in range(len(fieldNames)):
			if fieldValues[i].strip() == "":
				errmsg = errmsg + ('"%s" is a required field.\n\n' % fieldNames[i])
		if errmsg == "": 
			break # no problems found
		fieldValues = easygui.multenterbox(errmsg, title, fieldNames, fieldValues)
	time = parser.parse(fieldValues[0])
	
	msg = "Enter the home city of each of your participants"
	num_participants = int(fieldValues[1]) + 1
	fieldNames = ["Participant " + str(x) + " City" for x in range(1, num_participants)]
	fieldValues = []
	fieldValues = easygui.multenterbox(msg,title,fieldNames)
	while 1:
		if fieldValues == None: 
			break
		errmsg = ""
		for i in range(len(fieldNames)):
			if fieldValues[i].strip() == "":
				errmsg = errmsg + ('"%s" is a required field.\n\n' % fieldNames[i])
		if errmsg == "": 
			break # no problems found
		fieldValues = easygui.multenterbox(errmsg, title, fieldNames, fieldValues)
	meeting = {}
	participants = fieldValues
	msg = "Enter the salary or estimated salary of each of your participants"
	fieldNames = ["Participant " + str(x) + " Salary" for x in range(1, num_participants)]
	fieldValues = easygui.multenterbox(msg,title,fieldNames)
	for participant in range(0, len(participants)):
		geo = gmaps.geocode(participants[participant])
		location_table[participants[participant]] = {'lat':geo[0]["geometry"]["location"]["lat"],'lng':geo[0]["geometry"]["location"]["lng"]}
		meeting[participants[participant]] = int(fieldValues[participant])
	bml = best_meeting_location(time, meeting)
	msg = "Here are the best cities for your meeting, sorted by least potential loss:"
	choices = [str(l[0]) + " | Lost Productivity: " + str(l[1][0]) +" | Cost: " + str(l[1][1]) +" | Total Loss: " + str(l[1][2]) for l in bml]
	choice  = easygui.choicebox(msg, title, choices)
	return choice

output = load_workbook("output.xlsx")
gui()

# ws = output["SmallMeetingCost"]
	# sorted_city_prices = sorted(city_prices.items(), key=operator.itemgetter(1))
	# for x in range(0, len(sorted_city_prices)): 
	# 	ws['A' + str(x+1)] = sorted_city_prices[x][0]
	# 	ws['B' + str(x+1)] = sorted_city_prices[x][1]

	# c = load_workbook("cities.xlsx", read_only=True)
# ws = c["Cities"]
# for n in range(1,101):
# 	cities[ws.cell(column=3, row=n).value] = {'lat':ws.cell(column=5, row=n).value, 'lng':ws.cell(column=6, row=n).value}
# print cities

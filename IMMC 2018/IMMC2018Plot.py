import plotly.plotly as py
import plotly.figure_factory as ff
import plotly
import urllib
import json

plotly.tools.set_credentials_file(username='hwill12345', api_key='h2HOw9w2q4AKoSCsoIJg')

from openpyxl import load_workbook
import numpy as np
import pandas as pd

output = load_workbook("County.xlsx")
ws = output["CountyHospitals"]

df_sample = pd.read_csv('County.csv')

colorscale = ["#f7fbff","#ebf3fb","#deebf7","#d2e3f3","#c6dbef","#b3d2e9","#9ecae1",
              "#85bcdb","#6baed6","#57a0ce","#4292c6","#3082be","#2171b5","#1361a9",
              "#08519c","#0b4083","#08306b"]
endpts = list(np.linspace(0, 10, len(colorscale) - 1))
fips = df_sample['FIPS'].tolist()
values = df_sample['Quality Score'].tolist()

fig = ff.create_choropleth(
    fips=fips, values=values, scope=['usa'],
    binning_endpoints=endpts, colorscale=colorscale,
    show_state_data=False,
    show_hover=True, centroid_marker={'opacity': 0},
    asp=2.9, title='Hospital Quality Score by US County',
    legend_title='Hospital Quality Score'
)
print(py.plot(fig, filename='choropleth_full_usa'))


# def county_to_FIPS(state, county):
# 	search = urllib.parse.quote(f"{state}, {county.lower().capitalize()}")
# 	#http://coastwatch.pfeg.noaa.gov/erddap/convert/fipscounty.txt?county=CA%2C%20Monterey 
# 	url = f"http://coastwatch.pfeg.noaa.gov/erddap/convert/fipscounty.txt?county={search}"
# 	try:
# 		parsed_jsonID = json.loads(urllib.request.urlopen(url).read())
# 		return parsed_jsonID
# 	except:
# 		return "Failed"

# output = load_workbook("County.xlsx")
# ws = output["CountyHospitals"]
# ws["L1"] = "FIPS"
# for county in range(2, 1566):
# 	print(str(ws["K" + str(county)].value), str(ws["A" + str(county)].value))
# 	FIPS = county_to_FIPS(str(ws["K" + str(county)].value), str(ws["A" + str(county)].value))
# 	print(FIPS)
# 	if FIPS != "Failed":
# 		ws["L" + str(county)] = FIPS
# 		print("Written")
# output.save("County.xlsx")
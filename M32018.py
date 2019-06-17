# M3 2018 Model Calculation
# Team 10588

from openpyxl import load_workbook
from sklearn import linear_model, preprocessing
from SALib.sample import saltelli
from SALib.analyze import sobol
import numpy as np


ConsumerBehavior = load_workbook('ConsumerBehavior.xlsx', read_only=True)
#FoodConsumption = load_workbook('food_table1.xls', read_only=True)
StateFoodData = load_workbook('StateFoodData.xlsx', read_only=True)
FoodExpenditures = load_workbook('FoodExpenditures.xlsx', read_only=True)

def FoodNeeded(mealcost, low, medium, high, mode):
	pass

0.9
1.0
1.2
1.5
1.7
1.9
2.1
2.1
2.1

data = [[3768, 8383, 0.9, 0.4, 0.3], 
			[4437, 22167, 1.0, 0.5, 0.4],
			[5221, 34703, 1.2, 0.6, 0.5],
			[6028, 44589, 1.5, 0.4, 0.6],
			[6739, 59369, 1.7, 0.3, 0.6],
			[8436, 83595, 1.9, 0.3, 0.7],
			[10351, 120512, 2.1, 0.2, 0.8],
			[13550, 170704, 2.1, 0.2, 0.8],
			[16054, 345002, 2.1, 0.2, 0.9]]

test1 = [[8383, 0.9, 0.4, 0.3], 
			[22167, 1.0, 0.5, 0.4],
			[34703, 1.2, 0.6, 0.5],
			[44589, 1.5, 0.4, 0.6],
			[59369, 1.7, 0.3, 0.6],
			[83595, 1.9, 0.3, 0.7],
			[120512, 2.1, 0.2, 0.8],
			[170704, 2.1, 0.2, 0.8],
			[345002, 2.1, 0.2, 0.9]]

test2 = [[20500, 1,	0,	1],
		[135000, 2,	0,	2],
		[55000,	0,	2,	0],
		[45000,	1,	0,	0]]

#generate a model of polynomial features
#poly = preprocessing.PolynomialFeatures(degree=1)
#X_ = poly.fit_transform([[t[i] for i in range(1, 5)] for t in brackets])
#test1_ = poly.fit_transform(test1)
#test2_ = poly.fit_transform(test2)
X_ = [[t[i] for i in range(1, 5)] for t in data]

clf = linear_model.Ridge(alpha=float(input("alpha: ")))
#clf = linear_model.LinearRegression()
clf.fit(X_, [t[0] for t in data])

print(clf.score(X_, [t[0] for t in data]))
print(clf.intercept_)
print(clf.coef_)

print(clf.predict(test1))
print(clf.predict(test2))

print(f"({clf.intercept_} + {clf.coef_[0]}*B2 + {clf.coef_[1]}*D2 + {clf.coef_[2]}*E2 + {clf.coef_[3]}*F2)")

def P(X):
	return 2659.2483003294474 + 0.0325338098258917*X[0] + 1410.4836996101528*X[1] + -370.302803978762*X[2] + 432.74153542292726*X[3]

# Define the model inputs
problem = {
	'num_vars': 4,
	'names': ['I', 'A', 'S', 'C'],
	'bounds': [[0, 500000], [0,4], [0,3], [0,5]]
}

# Generate samples
param_values = saltelli.sample(problem, 100000, calc_second_order=True)

# Run model
Y = np.empty([param_values.shape[0]])

for i, X in enumerate(param_values):
    Y[i] = P(X)

# Perform analysis
Si = sobol.analyze(problem, Y, print_to_console=False)

# Print the first-order sensitivity indices
print("S1: ", Si['S1'], Si['S1_conf'])
print("ST: ", Si['ST'], Si['ST_conf'])


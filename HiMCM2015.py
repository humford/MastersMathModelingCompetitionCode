# HiMCM 2015 Problem B Model Calculation
# Team 5642

from openpyxl import load_workbook
wb = load_workbook('My_City_Crime_Data.xlsx')      
ws = wb.active

p_severity_beat_day_avg = 0.0016248907718
g_severity_beat_day_avg = 0.0023201624529
conviction_rate = 0.68
commuter_percentage = 0.06

def find_unique(col):
	unique = set()
	for row in range(2, 11163):
		unique.add(ws.cell(column=col, row=row).value)
	return unique

def crimes_beat(beat):
	crimes = {}
	for row in range(2, 11163): 
		if ws.cell(column=8, row=row).value == beat:
			location = ws.cell(column=5, row=row).value
			type_crime = ws.cell(column=3, row=row).value
			domestic = ws.cell(column=7, row=row).value
			arrest = ws.cell(column=6, row=row).value
			date = ws.cell(column=2, row=row).value.date()
			crimes[ws.cell(column=1, row=row).value] = [str(location), str(type_crime), str(domestic), arrest, date]
	return crimes

def crimes_beat_day(beat_crimes, day):
	day_crimes = {}
	for crime in beat_crimes:
		if beat_crimes[crime][4] == day:
			day_crimes[crime] = beat_crimes[crime][0:4]
	return day_crimes

def crimes_arrests_beat_day(crimes_beat_day):
	arrest_crimes = {}
	for crime in crimes_beat_day:
		if str(crimes_beat_day[crime][3]) == "Y":
			arrest_crimes[crime] = crimes_beat_day[crime]
	return arrest_crimes

def severity(crime):
	location = crime[0]
	domestic = crime[2]
	type_crime = crime[1]
	L = 0
	D = 0
	T = 0
	#Near home/Home 
	if location in ("APARTMENT", "RESIDENCE", "RESIDENCE PORCH/HALLWAY", "RESIDENCE-GARAGE", "NURSING HOME/RETIREMENT HOME", "DRIVEWAY - RESIDENTIAL", "CHA APARTMENT", "CHA PARKING LOT/GROUNDS", "PORCH", "COLLEGE/UNIVERSITY RESIDENCE HALL"):  
		L = float(5)/float(11)
	#Commercial buildings
	elif location in ("RESTAURANT", "CLEANING STORE", "SMALL RETAIL STORE", "RESIDENTIAL YARD (FRONT/BACK)", "BAR OR TAVERN", "GAS STATION", "MOVIE HOUSE/THEATER", "CONVENIENCE STORE", "TAVERN/LIQUOR STORE", "GROCERY FOOD STORE", "BARBERSHOP", "AUTO", "HOTEL/MOTEL", "COMMERCIAL / BUSINESS OFFICE", "WAREHOUSE", "DEPARTMENT STORE", "OTHER COMMERCIAL TRANSPORTATION", "VEHICLE-COMMERCIAL", "DRUG STORE", "ATHLETIC CLUB", "BANK", "CURRENCY EXCHANGE", "ATM (AUTOMATIC TELLER MACHINE)", "CAR WASH", "AIRPORT VENDING ESTABLISHMENT", "COIN OPERATED MACHINE", "PAWN SHOP", "APPLIANCE STORE"):  
		L = float(2)/float(11)
	#Public areas
	elif location in ("STREET", "CHURCH/SYNAGOGUE/PLACE OF WORSHIP", "SIDEWALK", "POLICE FACILITY/VEH PARKING LOT", "PARK PROPERTY", "SPORTS ARENA/STADIUM", "SCHOOL, PUBLIC, BUILDING", "CTA TRAIN", "AIRPORT BUILDING NON-TERMINAL - NON-SECURE AREA", "CTA STATION", "AIRPORT TERMINAL LOWER LEVEL - NON-SECURE AREA", "CTA BUS STOP", "AIRPORT TERMINAL UPPER LEVEL - SECURE AREA", "LAKEFRONT/WATERFRONT/RIVERBANK", "SCHOOL, PUBLIC, GROUNDS", "CTA PLATFORM", "CTA BUS", "COLLEGE/UNIVERSITY GROUNDS", "LIBRARY", "AIRPORT/AIRCRAFT", "AIRPORT BUILDING NON-TERMINAL - SECURE AREA", "BRIDGE", "HIGHWAY/EXPRESSWAY", "AIRPORT TERMINAL LOWER LEVEL - SECURE AREA", "AIRPORT PARKING LOT", "AIRPORT TRANSPORTATION SYSTEM (ATS)", "AIRPORT EXTERIOR - NON-SECURE AREA"):  
		L = float(3)/float(11)
	#Others
	elif location in ("PARKING LOT/GARAGE(NON.RESID.)", "OTHER", "ALLEY", "VEHICLE NON-COMMERCIAL", "VACANT LOT/LAND", "", "TAXICAB", "CHA HALLWAY/STAIRWELL/ELEVATOR", "HOSPITAL BUILDING/GROUNDS", "BOAT/WATERCRAFT", "ABANDONED BUILDING", "CTA GARAGE / OTHER PROPERTY", "VACANT LOT", "SCHOOL, PRIVATE, BUILDING", "GOVERNMENT BUILDING/PROPERTY", "CONSTRUCTION SITE", "MEDICAL/DENTAL OFFICE", "ANIMAL HOSPITAL", "JAIL / LOCK-UP FACILITY", "FACTORY/MANUFACTURING BUILDING", "FIRE STATION", "OTHER RAILROAD PROP / TRAIN DEPOT", "SCHOOL, PRIVATE, GROUNDS", "AIRPORT EXTERIOR - SECURE AREA", "POOL ROOM", "DAY CARE CENTER", "DELIVERY TRUCK", "AIRCRAFT"):  
		L = float(1)/float(11)
	if domestic == "Y":
		D = float(1)/float(4)
	elif domestic == "N":
		D = float(3)/float(4)
	if type_crime in ("CRIMINAL TRESPASS", "GAMBLING", "OTHER OFFENSE", "CRIMINAL", "INTERFERENCE WITH PUBLIC OFFICER", "PUBLIC PEACE VIOLATION"):
		T = float(1)/float(55)
	elif type_crime in ("CONCEALED CARRY LICENSE VIOLATION", "OTHER NARCOTIC VIOLATION"):
		T = float(2)/float(55)
	elif type_crime in ("NARCOTICS", "SEX OFFENSE", "PROSTITUTION", "LIQUOR LAW VIOLATION"):
		T = float(3)/float(55)
	elif type_crime in ("THEFT", "ASSAULT", "MOTOR VEHICLE THEFT", "WEAPONS VIOLATION", "DECEPTIVE PRACTICE", "INTIMIDATION"):
		T = float(4)/float(55)
	elif type_crime in ("OFFENSE INVOLVING CHILDREN", "CRIM SEXUAL ASSAULT"):
		T = float(5)/float(55)
	elif type_crime in ("STALKING", "ARSON", "BATTERY", "BURGLARY"):
		T = float(6)/float(55)
	elif type_crime in ("KIDNAPPING"):
		T = float(7)/float(55)
	elif type_crime in ("HOMICIDE", "ROBBERY"):
		T = float(8)/float(55)
	severity = float((30.0*L) + (10.0*D) + (70.0*T))
	return severity 



def crimes_severity_beat_day(crimes_beat_day):
	severity_beat_day = 0
	# severity_beat_day = sum(for crime in crimes_beat_day)
	for crime in crimes_beat_day:
		severity_beat_day += severity(crimes_beat_day[crime])
	return severity_beat_day

def population_beat(beat, sample_test):
	return float((9673.0 * commuter_percentage) + 9673.0)


def area_beat(beat, sample_test):
	return 7180.8

def avg_crimes_severity_beat_day(beats, day):
	pass

def safety_population_beat_day(beat, day):
	rb = crimes_beat(beat)
	rbd = crimes_beat_day(rb, day)
	rabd = crimes_arrests_beat_day(rbd)
	rsbd = crimes_severity_beat_day(rbd)
	rsabd = crimes_severity_beat_day(rabd)
	pbeat = population_beat(beat, True)

	if len(rbd)>0:
		beat_val = float(((rsbd/len(rbd))-((rsabd/len(rbd))*conviction_rate))/pbeat)
		r = (float(beat_val/p_severity_beat_day_avg))
		return r
	else:
		return 0


def safety_geography_beat_day(beat, day):
	rb = crimes_beat(beat)
	rbd = crimes_beat_day(rb, day)
	rabd = crimes_arrests_beat_day(rbd)
	rsbd = crimes_severity_beat_day(rbd)
	rsabd = crimes_severity_beat_day(rabd)
	gbeat = area_beat(beat, True)

	if len(rbd)>0:
		beat_val = float(((rsbd/len(rbd))-((rsabd/len(rbd))*conviction_rate))/gbeat)
		r = (float((beat_val/g_severity_beat_day_avg)))
		return r
	else:
		return 0


def safety_beat_day(beat, day):
	r = float((5*(safety_population_beat_day(beat,day))+2*(safety_geography_beat_day(beat, day)))/7) * 100.0
	return r

def safety_day(beats, day):
	l = []
	for beat in beats:
		l.append(safety_beat_day(beat, day))
	day_safety = float(sum(l))/len(l) if len(l) > 0 else float('nan')
	return day_safety

def safety(beats, days):
	l = []
	for day in days:
		l.append(safety_day(beats, day))
	return (float(sum(l))/len(l) if len(l) > 0 else float('nan'))

beats = find_unique(8)

days = set()
for row in range(2,11163):
	days.add(ws.cell(column=2, row=row).value.date())

print safety(beats, days)
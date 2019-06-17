from openpyxl import load_workbook
c = load_workbook("cities.xlsx", read_only=True)
ws = c["Cities"]
for n in range(1,101):
	cities[ws.cell(column=3, row=n).value] = {'lat':ws.cell(column=5, row=n).value, 'lng':ws.cell(column=6, row=n).value}
print cities